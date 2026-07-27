#!/usr/bin/env python3
"""Generate verified charge-service facts from the tracked business sources."""

from __future__ import annotations

import argparse
import copy
import hashlib
from pathlib import Path
from typing import Any

import openpyxl
import yaml
from docx import Document


ROOT = Path(__file__).resolve().parents[1]
SOURCE_DIR = (
    ROOT / "Workflow-China_charge_seriver-draft-9380" / "32-6.17（充电桩知识库）"
)
POLICY_PATH = ROOT / "shared" / "charge_service_policy.yaml"
OUTPUT_PATH = ROOT / "shared" / "charge_service.yaml"
MARKDOWN_PATH = (
    ROOT
    / "Workflow-China_charge_seriver-draft-9380"
    / "knowledge_bases"
    / "charge"
    / "generated_verified_facts.md"
)
SOURCE_FILES = (
    "常见问题解答.xlsx",
    "sys_menu_整理后.xlsx",
    "趋势云-标准充电桩功能清单V2.4.1-调整家充功能清单.xlsx",
    "Standard Operating Manual for Charging Piles.docx",
)


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _load_faq(path: Path) -> dict[str, str]:
    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    rows: dict[str, str] = {}
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows(values_only=True):
            values = [str(value).strip() if value is not None else "" for value in row]
            if len(values) >= 3 and values[1] and values[2]:
                rows[values[1]] = values[2]
    return rows


def _load_menu_paths(path: Path) -> set[str]:
    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    paths: set[str] = set()
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows(values_only=True):
            if row and row[0]:
                paths.add(str(row[0]).strip())
    return paths


def _load_feature_text(path: Path) -> list[str]:
    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    rows: list[str] = []
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows(values_only=True):
            text = " | ".join(str(value).strip() for value in row if value not in (None, ""))
            if text:
                rows.append(f"{sheet.title}: {text}")
    return rows


def _doc_paragraphs(path: Path) -> list[str]:
    return [paragraph.text.strip() for paragraph in Document(path).paragraphs]


def _require(mapping: dict[str, str], key: str) -> str:
    value = mapping.get(key)
    if not value:
        raise RuntimeError(f"required FAQ row not found: {key}")
    return value


def _require_line(lines: list[str], expected: str) -> str:
    for line in lines:
        if line == expected:
            return line
    raise RuntimeError(f"required source line not found: {expected}")


def _build_data() -> dict[str, Any]:
    policy = yaml.safe_load(POLICY_PATH.read_text(encoding="utf-8")) or {}
    faq = _load_faq(SOURCE_DIR / "常见问题解答.xlsx")
    menu_paths = _load_menu_paths(SOURCE_DIR / "sys_menu_整理后.xlsx")
    feature_rows = _load_feature_text(
        SOURCE_DIR / "趋势云-标准充电桩功能清单V2.4.1-调整家充功能清单.xlsx"
    )
    paragraphs = _doc_paragraphs(SOURCE_DIR / "Standard Operating Manual for Charging Piles.docx")

    pc_path = "充电桩>计费管理>充电计费模板"
    if pc_path not in menu_paths:
        raise RuntimeError(f"required menu path not found: {pc_path}")
    order_path = "财务>订单中心>充电桩订单>新能源车充电订单"
    if order_path not in menu_paths:
        raise RuntimeError(f"required menu path not found: {order_path}")
    _require_line(paragraphs, "Step 1: Click on the charging station ->Billing Management")
    _require_line(paragraphs, "Step 1: Click on the charging billing template ->Click on add template")
    _require_line(paragraphs, 'Step 2: Select "Time of use billing (peak and valley)", as shown in the following figure')
    _require_line(paragraphs, "Step 1: Enter My Venue ->Click on the Venue to enter the details page")
    _require_line(paragraphs, "Step 2: Click on billing settings ->select the desired template and confirm to save")
    fault_link = _require_line(paragraphs, "链接：/charge/pages/malfunction/malfunction").split("：", 1)[1]
    _require_line(paragraphs, "Step 1: Go to the configuration page ->Complete the form ->Submit and wait for platform review for offline processing")
    if not any("新建汽车充电计费模板" in row for row in feature_rows):
        raise RuntimeError("required V2.4.1 billing-template feature row not found")
    if not any("汽车充电订单导出" in row for row in feature_rows):
        raise RuntimeError("required V2.4.1 order-export feature row not found")

    sources = []
    combined = hashlib.sha256()
    for filename in SOURCE_FILES:
        path = SOURCE_DIR / filename
        digest = _sha256(path)
        combined.update(filename.encode("utf-8"))
        combined.update(digest.encode("ascii"))
        sources.append(
            {
                "path": str(path.relative_to(ROOT)),
                "sha256": digest,
            }
        )

    billing = {
        "standard_pc_path": pc_path.split(">"),
        "creation_steps_pc": [
            "进入充电桩 > 计费管理 > 充电计费模板",
            "点击添加模板",
            "选择分时计费（尖峰平谷）并按页面设置时段、电费和服务费",
        ],
        "association_steps_butler": [
            "场地审核通过后进入管家端“我的场地”",
            "选择场地进入详情页",
            "进入计费设置，选择所需模板并确认保存",
        ],
        "activation_rule": _require(faq, "计费模板创建后就直接生效吗？"),
        "replacement_rule": _require(faq, "站点更换计费模板后多久生效？"),
        "time_of_use_rule": _require(faq, "我们支持什么类型分时计费？"),
        "startup_balance_rule": _require(faq, "余额不足时可以启动充电吗？"),
        "guardrails": {
            "minimum_start_amount": "最低启动金额用于判断用户余额是否足以发起充电，不是创建或关联计费模板的前置条件。",
            "venue_audit": "管家端只有审核通过的场地才会显示在“我的场地”并可继续关联模板；这不是创建计费模板本身的前置条件。",
            "legacy_paths": [
                "IOT > 站点管理 > 计费模板（新）",
                "IOT > 故障报修",
            ],
        },
        "replies": {
            "location_zh": "PC后台普通汽车充电计费模板入口：充电桩 > 计费管理 > 充电计费模板。不同端或旧版本菜单可能不同，如你使用的不是PC后台，请补充端类型和当前页面。",
            "location_en": "On the PC admin portal, open Charging Pile > Billing Management > Charging Billing Template. Menus can differ by endpoint or legacy version, so the old IOT path should not be presented without that context.",
            "setup_zh": "PC后台：进入充电桩 > 计费管理 > 充电计费模板，点击添加模板，选择分时计费（尖峰平谷）并按页面设置时段、电费和服务费。模板创建后还需关联对应站点方可生效。最低启动金额是用户余额门槛，不是创建模板的前置条件。",
            "setup_en": "On the PC admin portal, open Charging Pile > Billing Management > Charging Billing Template, select Add Template, then configure time-of-use billing. A created template takes effect only after it is associated with the relevant station. The minimum start amount is a user-balance threshold, not a template-creation prerequisite.",
            "association_zh": "管家端：场地审核通过后，进入“我的场地” > 选择场地进入详情页 > 计费设置 > 选择所需模板 > 确认保存。最低启动金额是用户发起充电时的余额门槛，不是关联模板的前置条件。",
            "activation_zh": _require(faq, "计费模板创建后就直接生效吗？"),
            "replacement_zh": _require(faq, "站点更换计费模板后多久生效？"),
            "time_of_use_zh": _require(faq, "我们支持什么类型分时计费？"),
            "startup_balance_zh": _require(faq, "余额不足时可以启动充电吗？")
            + " 最低启动金额不是创建或关联计费模板的前置条件。",
            "guarded_zh": "关于计费模板，我只能按已核实资料回答：PC入口是“充电桩 > 计费管理 > 充电计费模板”；创建后需关联对应站点才生效；管家端在“我的场地 > 场地详情 > 计费设置”完成关联。最低启动金额是用户余额门槛，不是创建模板的前置条件。请说明你使用的是用户端、管家端还是PC后台，以及具体想确认哪一步。",
        },
        "evidence": [source["path"] for source in sources],
    }
    fault_repair = {
        "endpoint": "user",
        "custom_link": fault_link,
        "availability": "故障报修页需要由具体项目通过页面装修配置，是否开放以当前项目和应用版本为准。",
        "steps": ["进入已配置的故障报修页", "填写表单", "提交并等待平台审核后线下处理"],
        "replies": {
            "location_zh": f"用户端故障报修页需要由项目先完成页面装修配置；已配置时使用自定义链接 {fault_link}。进入后填写表单并提交，等待平台审核后线下处理。具体是否开放以当前项目和应用版本为准。"
        },
        "evidence": [
            "Workflow-China_charge_seriver-draft-9380/32-6.17（充电桩知识库）/Standard Operating Manual for Charging Piles.docx"
        ],
    }
    order_management = {
        "pc_path": order_path.split(">"),
        "filter_rule": _require(faq, "充电订单支持哪些筛选维度？"),
        "feature_rule": next(
            row for row in feature_rows if "国内PC端后台: 汽车充电订单导出" in row
        ),
        "replies": {
            "export_zh": "PC后台新能源车充电订单入口：财务 > 订单中心 > 充电桩订单 > 新能源车充电订单。已核实可按时间范围、站点、订单状态、充电起止时间等维度筛选，并导出当前查询结果；FAQ资料同时说明支持全量订单导出明细报表，具体以当前订单类型页面为准。",
            "export_en": "On the PC admin portal, open Finance > Order Center > Charging Pile Orders > New Energy Vehicle Charging Orders. The verified materials support filtering by time range, station, order status, and charging start/end time, then exporting the current result; the FAQ also records full-order detail export, subject to the current order type page.",
        },
        "evidence": [
            "Workflow-China_charge_seriver-draft-9380/32-6.17（充电桩知识库）/常见问题解答.xlsx",
            "Workflow-China_charge_seriver-draft-9380/32-6.17（充电桩知识库）/sys_menu_整理后.xlsx",
            "Workflow-China_charge_seriver-draft-9380/32-6.17（充电桩知识库）/趋势云-标准充电桩功能清单V2.4.1-调整家充功能清单.xlsx",
        ],
    }

    result = copy.deepcopy(policy)
    result["version"] = 2
    result["knowledge_revision"] = combined.hexdigest()[:16]
    result["generated_sources"] = sources
    result["verified_knowledge"] = {
        "billing_templates": billing,
        "user_fault_repair": fault_repair,
        "order_management": order_management,
    }
    return result


def _build_markdown(data: dict[str, Any]) -> str:
    billing = data["verified_knowledge"]["billing_templates"]
    fault = data["verified_knowledge"]["user_fault_repair"]
    order = data["verified_knowledge"]["order_management"]
    lines = [
        "# 充电客服已核实业务事实",
        "",
        f"> 知识版本: `{data['knowledge_revision']}`。本文件由真实 XLSX/DOCX 生成，不得手工修改。",
        "",
        "## 计费模板",
        "",
        f"- PC后台入口: {' > '.join(billing['standard_pc_path'])}",
        f"- 创建与生效: {billing['activation_rule']}",
        f"- 更换生效规则: {billing['replacement_rule']}",
        f"- 分时计费: {billing['time_of_use_rule']}",
        f"- 启动余额: {billing['startup_balance_rule']}",
        "- 管家端关联步骤: " + "；".join(billing["association_steps_butler"]) + "。",
        f"- 边界: {billing['guardrails']['minimum_start_amount']}",
        f"- 边界: {billing['guardrails']['venue_audit']}",
        "- 旧菜单边界: 未明确端类型或旧版上下文时，不输出 IOT 旧入口。",
        "",
        "## 用户端故障报修",
        "",
        f"- 自定义链接: `{fault['custom_link']}`",
        f"- 可用条件: {fault['availability']}",
        "- 操作步骤: " + "；".join(fault["steps"]) + "。",
        "- 边界: 不得把 PC 后台的 IOT 故障报修菜单回答为用户端入口。",
        "",
        "## 订单导出",
        "",
        f"- PC后台入口: {' > '.join(order['pc_path'])}",
        f"- 规则: {order['filter_rule']}",
        f"- 功能清单: {order['feature_rule']}",
        "- 边界: 端类型或订单类型不明确时，不把某一订单菜单扩展为所有订单入口。",
        "",
        "## 来源",
        "",
    ]
    lines.extend(
        f"- `{source['path']}` (`sha256:{source['sha256']}`)"
        for source in data["generated_sources"]
    )
    lines.append("")
    return "\n".join(lines)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--check", action="store_true", help="fail if generated files are stale")
    args = parser.parse_args()
    data = _build_data()
    yaml_text = yaml.safe_dump(data, allow_unicode=True, sort_keys=False, width=1000)
    markdown_text = _build_markdown(data)
    expected = ((OUTPUT_PATH, yaml_text), (MARKDOWN_PATH, markdown_text))
    if args.check:
        stale = [str(path.relative_to(ROOT)) for path, text in expected if not path.is_file() or path.read_text(encoding="utf-8") != text]
        if stale:
            raise SystemExit("generated charge facts are stale: " + ", ".join(stale))
        return
    for path, text in expected:
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(text, encoding="utf-8")


if __name__ == "__main__":
    main()
